"""Document parsers — return a list of (page_number, text) tuples.

Supported: PDF (pypdf), TXT/MD, XLSX (openpyxl), CSV. Scanned-image OCR
(Tesseract) is a documented extension point — not wired in the prototype.
"""
from pathlib import Path


def parse_file(path: str | Path) -> list[tuple[int, str]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _parse_pdf(path)
    if suffix in (".txt", ".md", ".log"):
        return _parse_text(path)
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    if suffix == ".csv":
        return _parse_text(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def _parse_pdf(path: Path) -> list[tuple[int, str]]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            pages.append((i, text))
    return pages


def _parse_text(path: Path) -> list[tuple[int, str]]:
    text = path.read_text(encoding="utf-8", errors="replace").strip()
    if not text:
        return []
    # Treat ~60 lines as one "page" so citations have page numbers.
    lines = text.splitlines()
    pages = []
    for i in range(0, len(lines), 60):
        chunk = "\n".join(lines[i : i + 60]).strip()
        if chunk:
            pages.append((i // 60 + 1, chunk))
    return pages


def _parse_xlsx(path: Path) -> list[tuple[int, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    pages = []
    for sheet_no, ws in enumerate(wb.worksheets, start=1):
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            pages.append((sheet_no, f"Sheet: {ws.title}\n" + "\n".join(rows)))
    return pages
